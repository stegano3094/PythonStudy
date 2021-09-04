# 엑셀 다루기

# 설치 : pip install openpyxl

# 시작 ========================================================
'''
import openpyxl  # openpyxl 라이브러리를 활용함

excel_file = openpyxl.load_workbook('test1.xlsx')
excel_sheet_list = excel_file.sheetnames  # 엑셀 시트들을 리스트로 가져옴
print(excel_sheet_list)

excel_sheet = excel_file[excel_sheet_list[0]]  # 가장 첫번째 시트를 가져옴
#excel_sheet = excel_file.active  # 처음으로 활성화 되어 있는 시트를 가져옴
print(excel_sheet)

for item in excel_sheet.rows:  # 각각의 행들을 추출
    print(item[0].value)
    print(item[1].value)
    print('') 

excel_file.close()  # 파일을 닫아줌
'''

# 모듈화 ======================================================
'''
import openpyxl  # openpyxl 라이브러리를 활용함

def readExcelFunc(excelFileName, sheetName):
    return_data = list()
    excel_file = openpyxl.load_workbook(excelFileName)

    if sheetName == '':
        excel_sheet = excel_file.active
    else:
        excel_sheet_list = excel_file.sheetnames  # 엑셀 시트들을 리스트로 가져옴
        excel_sheet = excel_file[excel_sheet_list[0]]  # 가장 첫번째 시트를 가져옴
        print(excel_sheet)
    
    for item in excel_sheet.rows:  # 각각의 행들을 추출
        data1 = list()
        for item2 in item:
            data1.append(item2.value)
        return_data.append(data1)

    excel_file.close()  # 파일을 닫아줌
    return return_data

data1 = readExcelFunc('test1.xlsx', '')

for item in data1:
    print(item[0])
    print(item[1])
'''


# 기능 익히기 =================================================
'''
import openpyxl

from random import *
print(int(random()*45)+1)  # 랜덤함수는 그냥 넣음

excel_file = openpyxl.Workbook()  # 새로운 엑셀 파일 생성

excel_sheet_active = excel_file.active  # 엑셀을 켤 때 첫 시트를 가져옴
excel_sheet_active.title = 'first'  # 엑셀 시트의 이름을 변경함

excel_sheet_active.append( ['a', 'b', 'c'] )  # 엑셀에 리스트 형식으로 입력
excel_sheet_active.append( ['', 'A', 'B', 'C'] )  # 2번 행의 셀에 자례대로 입력함

excel_sheet_active.column_dimensions['A'].width = 20  # 열 사이즈 변경
excel_sheet_active.column_dimensions['B'].width = 50

# https://openpyxl.readthedocs.io/en/stable/styles.html 다른 기능도 참고
# https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/borders.html
thin = openpyxl.styles.Side(border_style="thin", color='000000')  # border_style에 thin, thick 등 있음, 색상은 원하는대로
excel_sheet_active['B1'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)  # 주변 테두리 변경
excel_sheet_active['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')  # 셀 정렬 (horizontal= center, left, right)
excel_sheet_active['C1'].font = openpyxl.styles.Font(size=14, b=True, color="FF0000")  # 폰트 사이즈, 굵기, 색 변경
excel_sheet_active['A2'].fill = openpyxl.styles.PatternFill(start_color='00FF00', fill_type='solid')

excel_sheet_active.cell(row=3, column=5).font = openpyxl.styles.Font(size=14, b=True, color="FF0000")
excel_sheet_active.cell(row=3, column=5).hyperlink = "https://www.naver.com/"
excel_sheet_active.cell(row=3, column=5).alignment = openpyxl.styles.Alignment(horizontal='center')
excel_sheet_active.cell(row=3, column=5).fill = openpyxl.styles.PatternFill(start_color='00FF00', fill_type='solid')
excel_sheet_active.cell(row=3, column=5).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

listdata = [10, 20, 30, 40, 50, 60]
for index, item in enumerate(listdata):
    print("{0}, {1}".format(index, item))
    excel_sheet_active.append([item])

excel_file.save('test2.xlsx')  # 엑셀 파일을 저장
excel_file.close()  # 엑셀 파일 닫기
'''


# csv 파일 =================================================

# csv 파일은 pandas 라이브러리를 많이 사용하는데 구글링해서 사용할 것

import csv

# 쓰기
with open('test.csv', 'w') as writer_csv:
    writer = csv.writer(writer_csv, delimiter=',')

    writer.writerow(['i', 'can', 'do', 'it'])

    datas = ['abc', 'asdv', 'asdvdfv', 'asdgra']
    for item in datas:
        writer.writerow(item)
        print('write: {}'.format(item))

# open안에 encoding='utf-8-sig' => 저장된 데이터를 유니코드로 변환해서 읽어라는 의미
# 읽기
with open('test.csv', 'r') as reader_csv:
    datas = csv.reader(reader_csv, delimiter=',')

    for item in datas:
        print(item)

